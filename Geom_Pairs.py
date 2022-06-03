
# coding: utf-8

# # ** Import all modules and change working directory **
# 

# In[409]:


import os
import math
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.utils import get_column_letter
os.chdir("C:/Users/x01382250/Documents/Backtesteur Anthony")


# # ** Import the data **

# In[410]:


df_control = pd.read_excel("Backtester - Dispersion.xlsm",skiprows=7,sheet_name = "ControlPair", usecols= "G,H,I,J,K,L,M,N,P,R,T,U")

list_udlg = np.array(df_control.iloc[:,[0,1]])
vol_strikes = np.array(df_control.iloc[:,[2,3]])/100
basket_vol_strikes = np.array(df_control.iloc[:,4])/100
vega_not = np.array(df_control.iloc[:,[5,6]])
basket_vega_not = np.array(df_control.iloc[:,[7]])

N=int(df_control.iloc[0,8])
start_date = df_control.iloc[0,9]
start_date_range = df_control.iloc[0,10]
end_date_range = df_control.iloc[0,11]
df_data = pd.read_excel("Backtester - Dispersion.xlsm", sheet_name = "DataPairs", skiprows = 2, index_col= None).iloc[:,:4*len(list_udlg)]


# # ** Reformat tickers**

# In[411]:


def get_ticker(ticker):
    ticker = ticker.replace(" Equity","")
    ticker = ticker.replace(" Index","")
    return ticker


# #### We create a class "Vol Swap" containing all the data
#  * `__init__` : loads the parameters and compute in var terms for each stock and the basket: 
#      * $K_{var} = {K^2_{vol}}$
#      * $Not_{var} = \frac{100*Not_{vol}}{2K_{vol}}$
#  * `price vectors` : 
#      * from the data vector, we compute a vector containing all the backtest values from start date to today.
#      * compute the price; for each stock $j$, the payoff of the variance swap is:
#      
#     <center> $P_j = FRV_j^2 \wedge 2.5^2K_{var,\ j} - K^2_{var,\ j}$     &nbsp; where&nbsp;&nbsp;$FRV_j=\sqrt{\frac{252}{N}\sum_{t=1}^{N}{ln(\frac{S_{j , \ t}}{S_{j ,\ t-1}})}^2}$  </center>
#      
#      * for the basket, the $FRV$ is slighlty different; geometric computation gives arithmetic mean of log-returns:
#               
#   <center> $FRV_{basket}=\sqrt{\frac{252}{N}{\sum^{N}_{t=1}\left[\sum_{j=1}^{2}{ w_j*ln(\frac{S_{j , \ t}}{S_{j ,\ t-1}})}\right]^2}}$  </center>
#  
#      * we can compute the price of the Geometric dispersion :
#          <center> $P_{disp} = ((P_{S_1}+P_{S_2}-P_{basket}) \wedge 10*Not_{vol,basket})) \vee -10*Not_{vol,basket}))$ </center> 
#      which is basically the capped and floored long single stocks swaps and short basket swap.
#      
# 

# In[412]:


class var_swap:
    def __init__(self,N,data, vol_strike_s1,vol_strike_s2, vol_strike_basket,vega_not_s1,vega_not_s2, vega_not_basket): 
        self.N = N
        self.data = data 
        self.var_strike_s1 = vol_strike_s1**2
        self.var_strike_s2 = vol_strike_s2**2
        self.var_strike_b = vol_strike_basket**2
        
        self.var_not_s1 = vega_not_s1/(2*vol_strike_s1)
        self.var_not_s2 = vega_not_s2/(2*vol_strike_s2)
        self.var_not_b = vega_not_basket/(2*vol_strike_basket)              
        self.price_vector()
    def price_vector(self): 
        data_s1 = self.data.iloc[:,:2].set_index(list(self.data.iloc[:,[0]]))
        data_s1.index = pd.to_datetime(data_s1.index)
        
        data_s2 = self.data.iloc[:,2:].set_index(list(self.data.iloc[:,[2]]))
        data_s2.index = pd.to_datetime(data_s2.index)
                
        data_s1 = pd.DataFrame(pd.merge(data_s1,data_s2,right_index=True, left_index=True).iloc[:,0])
        data_s2= pd.DataFrame(pd.merge(data_s1,data_s2,right_index=True, left_index=True).iloc[:,1]) 

        data_s1 = np.log(1+data_s1.pct_change())
        data_s2 = np.log(1+data_s2.pct_change())
        
        
        data_basket = pd.DataFrame()
        data_basket["Basket"] = (0.5*data_s1.iloc[:,0]+0.5*data_s2.iloc[:,0])**2 
        data_s1 = data_s1**2
        data_s2 = data_s2**2
        
        data_s1.iloc[:,:1] =(data_s1.iloc[:,:1]).rolling(N).mean()*252 
        data_s1.iloc[:,:1]=((((data_s1.iloc[:,:1])).clip(upper=self.var_strike_s1*2.5**2))-self.var_strike_s1) * self.var_not_s1*100
        data_s1 = data_s1.iloc[N:,:]
        
        data_s2.iloc[:,:1] =(data_s2.iloc[:,:1]).rolling(N).mean()*252 
        data_s2.iloc[:,:1]=((((data_s2.iloc[:,:1])).clip(upper=self.var_strike_s2*2.5**2))-self.var_strike_s2)*self.var_not_s2*100
        data_s2 = data_s2.iloc[N:,:]
        
        data_basket.iloc[:,:1] = ((data_basket.iloc[:,:1])).rolling(N).mean()*252
        data_basket.iloc[:,:1]=((((data_basket.iloc[:,:1])).clip(upper=self.var_strike_b*2.5**2))-self.var_strike_b)*self.var_not_b*100
        data_basket = data_basket.iloc[N:,:]

        self.data = pd.merge(data_s1,data_s2,left_index = True, right_index = True)
        self.data = pd.merge(self.data,data_basket,left_index=True,right_index=True)
        self.data.columns = [get_ticker(self.data.columns[0]),get_ticker(self.data.columns[1]),"Basket " + get_ticker(self.data.columns[0]) + " & " + get_ticker(self.data.columns[1])]


# ** Compute the backtest for all the stocks **
# * we compute the backtest and convert the results in USD to vol points, to store both data.
# * for the geometric backtest, we work the dates by taking the common calendars of the two stocks. To display all pairs, we use the common calendars (it should not vary much, as the pairs are generally on the same calendar).

# In[413]:


backtest=pd.DataFrame({})
final_only = pd.DataFrame({})
usd_pnl= pd.DataFrame({})

for i in range(len(df_data.columns)//4):

    data = df_data.iloc[:,4*i:4*(i+1)]

    vol_K_stocks = vol_strikes[i]
    vol_K_basket = basket_vol_strikes[i]
    vol_not_stocks = vega_not[i]
    vol_not_basket = basket_vega_not[i]
    swap = var_swap(N,data,vol_K_stocks[0],vol_K_stocks[1],vol_K_basket,vol_not_stocks[0], vol_not_stocks[1],vol_not_basket)
    result = swap.data

    col_name = "Final " + get_ticker(result.columns[0]) + " & " + get_ticker(result.columns[1])
    result[col_name] = (result.iloc[:,0] + result.iloc[:,1]- result.iloc[:,2]) #cest en usd'
    usd_pnl[col_name] = result.iloc[:,3].clip(upper = 10*float(vol_not_basket),lower= -10*float(vol_not_basket))
    result.iloc[:,3] =result.iloc[:,3].clip(upper = 10*float(vol_not_basket),lower= -10*float(vol_not_basket))/basket_vega_not[i]
    if i==0:
        backtest= result
        final_only = result[[col_name]]
    else :
        backtest = pd.merge(backtest, result, left_index=True, right_index=True,sort=False)
        final_only = pd.merge(final_only,result[[col_name]], left_index=True, right_index=True,sort=False)

            
backtest.index.names= ["Dates"]
final_only.index.names= ["Dates"]

backtest = backtest.fillna(0)
final_only = final_only.fillna(0)
usd_pnl = usd_pnl.fillna(0)
pnl_points =  usd_pnl.sum(axis=1)/(((basket_vega_not.T)*((usd_pnl)!=0)).sum(axis=1))
final_only["Dollar P&L of the sum"] = usd_pnl.fillna(0).sum(axis = 1)
final_only["Vega points P&L of the sum"] = pnl_points
final_only["Percentage of Baskets"] = ((final_only.iloc[:,:-2])!=0).sum(axis=1)/(len(final_only.columns)-2)*100


# # Compute Hit Ratios, means,...

# In[414]:


def compute_dataframes_ratio(backtest,start,end):
    backtest_compute = backtest.iloc[backtest.index.get_loc(start, method='nearest'):backtest.index.get_loc(end, method='nearest')+1,:-1].replace(0,np.NaN)
    results = pd.DataFrame(columns = backtest_compute.columns)
    results.loc["Mean"] = backtest_compute.mean()
    results.loc["Min"] = backtest_compute.min()
    results.loc["Max"] = backtest_compute.max()
    results.loc["Last"] = backtest_compute.iloc[-1,:]
    results.loc["Hit Ratio"] = ((backtest_compute.replace(np.NaN,0))>0).sum()/(((backtest_compute.replace(np.NaN,0))!=0).sum())
    results.loc["Hit Ratio"]=((results.loc['Hit Ratio']).apply('{:.00%}'.format))
    results.loc["First valid date"] = backtest_compute.replace(0,np.NaN).apply(lambda series: series.first_valid_index())
    return results


# ** Compute the data from backtest start date to end date **

# In[415]:


start = backtest.index[0]
end  =backtest.index[-1]
summary_no_range = compute_dataframes_ratio(final_only,start,end)


# ** Export the data to excel and format **

# In[416]:


excel_file = 'Geometric Pairs Backtest - Results.xlsx'
writer = pd.ExcelWriter(excel_file, engine='xlsxwriter',datetime_format='dd/mm/yyyy')
backtest.to_excel(writer,sheet_name="Prices (all)")


final_only.to_excel(writer, sheet_name="Prices (only finals)")
summary_no_range.to_excel(writer,sheet_name = "Summary")

workbook = writer.book
fmt_bleu= workbook.add_format({"bg_color": "#DAEEF3"})
worksheet= workbook.add_worksheet("Graph")
worksheet.set_column('A:ZZ', None, fmt_bleu)

percentage = workbook.add_format({'num_format': '0.00%'})
writer.save()


# ** Create graphics : each one plots 2 pairs. ** 

# In[418]:


fileSavePath = "C:/Users/x01382250/Documents/Backtesteur Anthony/"
filename = "Geometric Pairs Backtest - Results.xlsx"

wb = load_workbook(fileSavePath + filename)
wb.active = 3 
worksheet = wb.active

wb.active = 3
worksheet = wb.active

nb_pairs = 24
for i in range(max(1,len(final_only.columns)//2-1)):
    
    maxcol = min(2*(i+1)+1,len(final_only.columns)+1) if i!= 0 else 2
    source=wb["Prices (only finals)"]
    chart = LineChart()
    data = Reference(wb["Prices (only finals)"], min_col = 2+2*i,min_row = 1, max_row = len(final_only.index),max_col =maxcol) 
    category = Reference(wb["Prices (only finals)"], min_col = 1,min_row = 2, max_row = len(final_only.index),max_col =1) 

    chart.add_data(data,titles_from_data=True)
    chart.set_categories(category)
    chart.x_axis.title = "Dates"
    chart.y_axis.title = "P&L in Vol points"
    chart.title = "Backtest Pairs " + str(2*i+1) + " to " + str(2*(i+1))
    props = GraphicalProperties(solidFill="E9F1F5") 
    chart.plot_area.graphicalProperties = props

    chart.x_axis.number_format = 'yyyy'
    chart.x_axis.majorUnit = 365.25
    
    
    chart.y_axis.scaling.min = -11
    chart.y_axis.scaling.max = 11

    chart.width =20.37
    col = get_column_letter((i%3)*12+1)
    cell = col + str(3+14*(i//3))
    for ind in range(len(chart.series)):
        chart.series[ind].graphicalProperties.line.width = 15000
    
    worksheet.add_chart(chart, cell)


# ** Add the Dollar P&L and the percentage of baskets **

# In[419]:


wb.active = 2
worksheet = wb.active

source=wb["Prices (only finals)"]
chart = LineChart()

data = Reference(wb["Prices (only finals)"], min_col = len(final_only.columns)-1,min_row = 1, max_row = len(final_only.index)+1,max_col =len(final_only.columns)-1) 
category = Reference(wb["Prices (only finals)"], min_col = 1,min_row = 1, max_row = len(final_only.index)+1,max_col =1) 
chart.add_data(data,titles_from_data=True)
chart.set_categories(category)
chart.x_axis.title = "Dates"
chart.y_axis.title = "P&L in USD"
chart.title = "Backtest Pairs Total USD P&L"
chart.x_axis.majorUnit = 365.25

chart2 = LineChart()
data2 = Reference(wb["Prices (only finals)"], min_col = len(final_only.columns)+1,min_row = 1, max_row = len(final_only.index)+1,max_col =len(final_only.columns)+1) 
chart2.add_data(data2,titles_from_data=True)
chart2.x_axis.title = "Dats"
chart2.y_axis.title = "Percentage of Baskets"
chart2.title = "Backtest Pairs Total USD P&L"
chart2.y_axis.axId = 200
chart2.y_axis.crosses = 'max' 
chart2.x_axis.majorUnit = 365.25
chart2.x_axis.number_format = 'yyyy'
chart2.y_axis.majorGridlines = None

chart += chart2

props = GraphicalProperties(solidFill="E9F1F5") 
chart.plot_area.graphicalProperties = props
chart.width =40.5
chart.height=10.5
chart.x_axis.number_format = 'yyyy'
chart.x_axis.majorUnit = 365.25
worksheet.add_chart(chart,"A10")


# ** Add the Vega points P&L and the percentage of baskets **

# In[420]:


worksheet = wb.active

chart = LineChart()

data = Reference(wb["Prices (only finals)"], min_col = len(final_only.columns),min_row = 1, max_row = len(final_only.index)+1,max_col =len(final_only.columns)) 
category = Reference(wb["Prices (only finals)"], min_col = 1,min_row = 1, max_row = len(final_only.index)+1,max_col =1) 
chart.add_data(data,titles_from_data=True)
chart.set_categories(category)
chart.x_axis.title = "Dates"
chart.y_axis.title = "P&L in Vega points"
chart.title = "Backtest Pairs Total Vega points P&L"
chart.x_axis.majorUnit = 365.25

chart2 = LineChart()
data2 = Reference(wb["Prices (only finals)"], min_col = len(final_only.columns)+1,min_row = 1, max_row = len(final_only.index)+1,max_col =len(final_only.columns)+1) 
chart2.add_data(data2,titles_from_data=True)
chart2.x_axis.title = "Dats"
chart2.y_axis.title = "Percentage of Baskets"
chart2.title = "Backtest Pairs Total Vega points P&L"
chart2.y_axis.axId = 200
chart2.y_axis.crosses = 'max' 
chart2.x_axis.majorUnit = 365.25
chart2.x_axis.number_format = 'yyyy'
chart2.y_axis.majorGridlines = None

chart += chart2

props = GraphicalProperties(solidFill="E9F1F5") 
chart.plot_area.graphicalProperties = props
chart.width =40.5
chart.height=10.5
chart.x_axis.number_format = 'yyyy'
chart.x_axis.majorUnit = 365.25


worksheet.add_chart(chart,"Y10")


# ** Add Vega P&L without weights **

# In[421]:


worksheet = wb.active

chart = LineChart()

data = Reference(wb["Prices (only finals)"], min_col = len(final_only.columns),min_row = 1, max_row = len(final_only.index)+1,max_col =len(final_only.columns)) 
category = Reference(wb["Prices (only finals)"], min_col = 1,min_row = 1, max_row = len(final_only.index)+1,max_col =1) 
chart.add_data(data,titles_from_data=True)
chart.set_categories(category)
chart.x_axis.title = "Dates"
chart.y_axis.title = "P&L in Vega points"
chart.title = "Backtest Pairs Total Vega points P&L"
chart.x_axis.majorUnit = 365.25


props = GraphicalProperties(solidFill="E9F1F5") 
chart.plot_area.graphicalProperties = props
chart.width =40.5
chart.height=10.5
chart.x_axis.number_format = 'yyyy'
chart.x_axis.majorUnit = 365.25

worksheet.add_chart(chart,"Y31")


# ** Add a recap worksheet **

# In[422]:


wb.create_sheet("Input Recap")
wb.active  = 4
worksheet = wb.active


worksheet["A1"].offset(0,0).value = str("Underlying #1")
worksheet["A1"].offset(0,1).value = "Underlying #2"
worksheet["A1"].offset(0,2).value = "Underlying #1 VolStrike"
worksheet["A1"].offset(0,3).value = "Underlying #2 VolStrike"
worksheet["A1"].offset(0,4).value = "Basket VolStrike"
worksheet["A1"].offset(0,5).value = "Underlying #1 Vega Notional"
worksheet["A1"].offset(0,6).value = "Underlying #2 Vega Notional"
worksheet["A1"].offset(0,7).value = "Basket Vega Notional"

for i in range(1,len(list_udlg)+1):
        worksheet["A1"].offset(i,0).value = list_udlg[i-1][0]
        worksheet["A1"].offset(i,1).value = list_udlg[i-1][1]
        worksheet["A1"].offset(i,2).value = vol_strikes[i-1][0]
        worksheet["A1"].offset(i,3).value = vol_strikes[i-1][1]
        worksheet["A1"].offset(i,4).value = basket_vol_strikes[i-1]
        worksheet["A1"].offset(i,5).value = vega_not[i-1][0]
        worksheet["A1"].offset(i,6).value = vega_not[i-1][1]
        worksheet["A1"].offset(i,7).value = float(basket_vega_not[i-1])
    
worksheet["A1"].offset(0,8).value = "Nexp"
worksheet["A1"].offset(1,8).value = N
worksheet["A1"].offset(0,9).value = "Backtest Start Date"
worksheet["A1"].offset(1,9).value = start_date

worksheet["A1"].offset(0,10).value = "Range Start Date"
worksheet["A1"].offset(1,10).value = start_date_range
worksheet["A1"].offset(0,11).value = "Range End Date"
worksheet["A1"].offset(1,11).value = end_date_range


wb.save(fileSavePath + filename)

