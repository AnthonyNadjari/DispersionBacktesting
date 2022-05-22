#!/usr/bin/env python
# coding: utf-8

# In[75]:



import os
import math
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
os.chdir("C:/Users/nadja/OneDrive/Bureau/code")


df_data = pd.read_excel("Backtester - Dispersion.xlsm", sheet_name = "Data", skiprows = 1, index_col= None).dropna(axis=1, how="all")
df_control = pd.read_excel("Backtester - Dispersion.xlsm",skiprows=6,sheet_name = "Control", usecols= "G,I,K,M,O,Q,S,U")

data_index = df_data.iloc[:,[1]].set_index(np.array(df_data.iloc[:,0]))
data_index = data_index.loc[data_index.index.notnull()]
N=df_control.iloc[2,3]
strike=np.array(df_control.iloc[2:,4])
start_date = df_control.iloc[2,6]
strike_index = df_control.iloc[2,5]
weights = np.array(df_control.iloc[1:,1].dropna(axis = 0))

weights = np.array(df_control.iloc[1:,1].dropna(axis = 0))
tickers_name =  np.array(df_control.iloc[1:,0].dropna(axis = 0))
dico={}
for key,weight in zip (tickers_name, weights):
    dico[key]=weight
dic_weights={}
dic_weights = {dico[key] for key in sorted(dico)}
sorted_dict = {key: value for key, value in sorted(dico.items())}
weights = np.array(list(sorted_dict.values()))

 
class vol_swap:
    def __init__(self,N,strike,data):
        self.N = N
        self.K = strike
        self.data = data 
        self.price_vector()
    def price_vector(self):
        ret  = self.data.pct_change()**2
        ret.iloc[:,:1] =np.sqrt((ret.iloc[:,:1]).rolling(N).mean()*252)
        ret.iloc[:,:1]=(ret.iloc[:,:1].clip(upper=self.K*2.5)-self.K )*100
        self.prices = ret

def compute_backtest():
    backtest = df_data.iloc[:,:2]
    backtest=backtest.set_index(np.array(backtest.iloc[:,0])).iloc[:,:1]
    for i in range(1,int(len(df_data.columns)/2)):  
        data = df_data.iloc[:, [2*i,2*i+1]]
        data = data.set_index(np.array(data.iloc[:,0]))
        data = data.iloc[:,1:]
        swap_stock = vol_swap(N,strike[i-1],data)
        swap_stock.prices  = swap_stock.prices[swap_stock.prices.index.notnull()]
        if i==1:
            backtest= swap_stock.prices.iloc[N:,:]
        else :
            backtest = pd.merge(backtest, swap_stock.prices.iloc[N:,:], left_index=True, right_index=True)
    backtest = backtest.reindex(sorted(backtest.columns), axis=1)
    return backtest
backtest = compute_backtest()

backtest=backtest.fillna(0)
backtest["weighted_sum"] = (weights.T*backtest).sum(axis=1)
df_is_na = backtest.iloc[:,:-1]!=0
backtest["sum_of_weights"] =(weights.T*df_is_na).sum(axis = 1)
backtest["result_stocks"] = backtest["weighted_sum"]/backtest["sum_of_weights"]
index_swap = vol_swap(N,strike_index,data_index)
backtest =pd.merge(backtest,index_swap.prices.iloc[N:,:],left_index=True,right_index=True) 
backtest["result"] = backtest["result_stocks"] -  backtest.iloc[:,-1]
N=60
backtest_60 = compute_backtest()
backtest_60=backtest_60.fillna(0)
backtest_60["weighted_sum"] = (weights.T*backtest_60).sum(axis=1)
df_is_na_60 = backtest_60.iloc[:,:-1]!=0
backtest_60["sum_of_weights"] =(weights.T*df_is_na_60).sum(axis = 1)
backtest_60["result_stocks"] = backtest_60["weighted_sum"]/backtest_60["sum_of_weights"]
index_swap_60 = vol_swap(N,strike_index,data_index)
backtest_60 =pd.merge(backtest_60,index_swap_60.prices.iloc[N:,:],left_index=True,right_index=True) 
backtest_60["result"] = backtest_60["result_stocks"] -  backtest_60.iloc[:,-1]
def compute_dataframes_ratio(backtest,start,end):
    backtest_compute = backtest.drop(["weighted_sum","sum_of_weights","result_stocks","result"], axis = 1).iloc[backtest.index.get_loc(start, method='nearest'):backtest.index.get_loc(end, method='nearest')+1,:]
    backtest_compute = backtest_compute.sub(backtest_compute.iloc[:,-1],axis=0).iloc[:,:-1]
    results = pd.DataFrame(columns = backtest_compute.columns)
    results.loc["Mean"] = backtest_compute.mean()
    results.loc["Min"] = backtest_compute.min()
    results.loc["Max"] = backtest_compute.max()
    results.loc["Last"] = backtest_compute.iloc[-1,:]
    results.loc["Hit Ratio"] = (backtest_compute>0).sum()/len(backtest.index)
    results.loc["Hit Ratio"]=((results.loc['Hit Ratio']).apply('{:.00%}'.format))
    return results
start = df_control.iloc[2,7]
end = df_control.iloc[3,7]
results_range = compute_dataframes_ratio(backtest,start,end)
results_no_range = compute_dataframes_ratio(backtest,start_date,backtest.index[-1])
results_range.loc["Last (60 days)"] = backtest_60.iloc[backtest_60.index.get_loc(end, method='nearest'),:]
results_no_range.loc["Last (60 days)"]  = backtest_60.iloc[-1,:]


# In[76]:


excel_file = 'Backtest - Results.xlsx'
sheet_name = 'Prices'
writer = pd.ExcelWriter(excel_file, engine='xlsxwriter',datetime_format='dd/mm/yyyy')
backtest.to_excel(writer, sheet_name=sheet_name)
backtest_60.to_excel(writer,sheet_name="Prices 60")

results_no_range.to_excel(writer,sheet_name="Mono")
results_range.to_excel(writer,sheet_name="Mono",startrow = 15)


# In[77]:


workbook = writer.book
fmt_bleu= workbook.add_format({"bg_color": "#DAEEF3"})
worksheet= workbook.add_worksheet("Graph")
worksheet.set_column('A:ZZ', None, fmt_bleu)

percentage = workbook.add_format({'num_format': '0.00%'})




chart1 = workbook.add_chart({'type': 'line'})

chart1.add_series({
    'categories': ['Prices', 1, 0, len(backtest.index), 0],
    'values':     ['Prices', 1, len(backtest.columns), len(backtest.index), len(backtest.columns)],
})

chart1.set_plotarea({
    'fill':   {'color': '#E9F1F5'},
    'line':   {'color' : "#46AAC5"}
})


chart1.set_title ({'name': 'Backtest'})
chart1.set_x_axis({
    'name': 'Dates',
    'visible': True,
    'line': {'width': 1.25, 'dash_type': 'dash'}, 
    "date_axis": True, 
    'min':backtest.index[0] ,
    'max': backtest.index[len(backtest.index)-1],})

chart1.set_y_axis({'name': 'P&L in Vol points'})
chart1.set_legend({'none': True})
worksheet.insert_chart('D2', chart1, {'x_offset': 25, 'y_offset': 10,'x_scale': 2, 'y_scale': 2})

worksheet60= workbook.add_worksheet("Graph 60")
worksheet60.set_column('A:ZZ', None, fmt_bleu)



writer.save()
writer.close()


# In[111]:



from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.utils import get_column_letter


fileSavePath = "C:/Users/nadja/OneDrive/Bureau/code/"
filename = "Backtest - Results.xlsx"
newSheetName = 'test'
sheet = newSheetName

wb = load_workbook(fileSavePath + filename)
print(wb.sheetnames)
wb.active = 3
worksheet = wb.active

nb_pairs = 24
for i in range(0,9):
    j=i//4
    print(j)
    wb.active = 3
    worksheet = wb.active
    source=wb["Prices"]
    chart = LineChart()
    data = Reference(wb["Prices"], min_col = 2+3*i,min_row = 2, max_row = len(backtest.index),max_col =3*i+4) 
    category = Reference(wb["Prices"], min_col = 1,min_row = 2, max_row = len(backtest.index),max_col =1) 

    chart.add_data(data)# titles_from_data=True)
    chart.set_categories(category)
    chart.x_axis.title = "Dates"
    chart.y_axis.title = "P&L in Vol points"
    chart.title = "Backtest Pairs" + str(3*i+1) + " to " + str(3*(i+1))
    props = GraphicalProperties(solidFill="E9F1F5") 
    chart.plot_area.graphicalProperties = props

    chart.x_axis.number_format = 'yyyy'
    chart.x_axis.majorUnit = 365.25
    chart.width =13.58
    col = get_column_letter((i%4)*8+1)
    cell = col + str(3+14*j)
    
    
  
    worksheet.add_chart(chart, cell)
    
wb.save(fileSavePath + filename)


# In[104]:


from openpyxl.utils import get_column_letter
print(get_column_letter(1)+ str(1))
print(len(backtest.columns))
print(13//4)

